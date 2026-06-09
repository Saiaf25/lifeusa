"""Build LifeUSA Technical SEO Master Sheet from Makeen template.

Adapts the TMD audit playbook to Lifeusa.org (charity). Headline finding:
subdomain governance + Wix content-language header bug, not on-page hygiene.

Inputs:  /Users/saiaf/Downloads/TMD/Mishraq/Technical SEO [audit template] previous example for makeen.xlsx
Outputs: /Users/saiaf/Downloads/lifeusaorg/LifeUSA Technical SEO [Master Sheet].xlsx
"""
import shutil, openpyxl, os
from openpyxl.styles import Font, PatternFill, Alignment

SRC = '/Users/saiaf/Downloads/TMD/Mishraq/Technical SEO [audit template] previous example for makeen.xlsx'
DST = '/Users/saiaf/Downloads/lifeusaorg/LifeUSA Technical SEO [Master Sheet].xlsx'

shutil.copy2(SRC, DST)
wb = openpyxl.load_workbook(DST)


def safe_set(ws, row, col, value):
    """Set a cell value; unmerge if needed; clear hyperlink."""
    cell = ws.cell(row=row, column=col)
    merged_to_remove = []
    for mrange in list(ws.merged_cells.ranges):
        if (mrange.min_row <= row <= mrange.max_row and
                mrange.min_col <= col <= mrange.max_col):
            merged_to_remove.append(str(mrange))
    for mrange in merged_to_remove:
        ws.unmerge_cells(mrange)
    cell = ws.cell(row=row, column=col)
    cell.value = value
    cell.hyperlink = None
    return cell


def clear_row(ws, row, col_start=1, col_end=12):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        for mrange in list(ws.merged_cells.ranges):
            if (mrange.min_row <= row <= mrange.max_row and
                    mrange.min_col <= c <= mrange.max_col):
                ws.unmerge_cells(str(mrange))
        cell.value = None
        cell.hyperlink = None


def clear_sheet_data(ws, header_row=1, max_clear_row=200):
    for r in range(header_row + 1, max_clear_row + 1):
        clear_row(ws, r, 1, 12)


# ============================================================================
# TAB 1: Executive Summary
# ============================================================================
ws = wb['Executive Summary']
for r in range(1, 60):
    clear_row(ws, r, 1, 12)

EXEC = [
    (1,  1, "LIFEUSA.ORG - Technical SEO Audit Summary"),
    (2,  1, "Website: https://www.lifeusa.org/"),
    (3,  1, "Audit Date: May 2026"),
    (4,  1, "Project owner: Saiaf Gamal (freelance consultation)"),
    (5,  1, "Client: Life for Relief and Development (charity, established 1992, Southfield MI USA)"),
    (7,  1, "ISSUE SUMMARY BY PRIORITY"),
    (8,  1, "Priority"), (8, 2, "Description"),
    (9,  1, "Critical"),  (9, 2, "Immediate attention required - blocking SEO performance and donor visibility"),
    (10, 1, "High"),      (10, 2, "Important issues affecting rankings and visibility"),
    (11, 1, "Medium"),    (11, 2, "Moderate impact - should be addressed in short term"),
    (12, 1, "Low"),       (12, 2, "Minor optimizations for ongoing improvement"),
    (14, 1, "KEY FINDINGS"),
    (15, 1, "CRITICAL: Subdomain governance is broken. arabic.lifeusa.org does not resolve in DNS (1 referring domain wasted). donation.lifeusa.org runs a second donations platform (Givecloud) that duplicates donate.lifeusa.org (TouchPoints), splitting donor data. staff.lifeusa.org is a publicly reachable WordPress install with its own robots.txt and risks indexation of internal content."),
    (16, 1, "CRITICAL: HTTP `content-language` header is wildly incorrect across pages. Homepage returns `zh-CN` (Chinese), /about returns `tr-TR` (Turkish), /gaza returns `fr-FR` (French), /ar returns `en` (English). The HTML <html lang> attribute is correct, but Wix is sending mismatched HTTP language headers per request. This is a Wix multilingual misconfiguration that confuses Google's language signals and likely contributes directly to the brand's English-search visibility problem."),
    (17, 1, "CRITICAL: Hreflang locale `ar-jo` declared for the Arabic version. For a US-headquartered global Muslim charity, the audience is global Arabic speakers, not specifically Jordanian. Should be `ar` (default) or a more representative locale."),
    (18, 1, "HIGH: 122 URLs (35%) carry a noindex directive. This needs an urgent audit. Review of Screaming Frog export shows likely accidental noindex on key program/landing pages."),
    (19, 1, "HIGH: 90% of images missing width/height attributes (1,035 images). Major Cumulative Layout Shift (CLS) risk. Wix template-level fix."),
    (20, 1, "HIGH: 47% of images over 100KB (543 images). Page-speed and LCP risk, especially on Gaza, Programs, and donation landing pages."),
    (21, 1, "HIGH: Schema markup is generic (WebSite + ProfessionalService). For a registered nonprofit, the more appropriate type is `NGO` (subclass of Organization) plus DonateAction. Article schema on 441 blog posts entirely missing."),
    (22, 1, "HIGH: Donor-trust security headers absent. CSP missing on 89% of URLs, X-Frame-Options on 89%, Referrer-Policy on 93%. These do not affect ranking directly but matter for donor trust signals and protection against clickjacking on donation flows."),
    (23, 1, "MEDIUM: 144 pages (54%) missing meta descriptions. 97 page titles (37%) identical to H1. 72 titles (27%) over 60 characters. Wix template defaults. Content-team fix."),
    (24, 1, "MEDIUM: 451 images missing alt text + 20 missing alt attribute. Directly maps to client's pain that 'images attached to Arabic articles do not appear in English search.' Without alt text, no language search will surface them."),
    (25, 1, "MEDIUM: 7 exact-duplicate pages, 12 duplicate meta descriptions, 15 missing canonicals."),
    (26, 1, "MEDIUM: Sitemap contains template leftover URLs: /no, /copy-of-freewill-landing-page, /b2s, /fb-intl. Likely Wix scaffolding never cleaned up."),
    (27, 1, "STRENGTH: Apex 301 to www works correctly (Wix Pepyaka, HSTS enabled). The Ahrefs 'apex 0 traffic' is a reporting artifact, not a broken redirect."),
    (28, 1, "STRENGTH: 441 indexed blog posts and 33 ranking organic pages with 1,069 monthly organic traffic. The donate. subdomain has 214 referring domains and a hardened security posture (CSP, HSTS, CF). Strong base to keep building on."),
    (29, 1, "STRENGTH: hreflang implementation exists (x-default + ar-jo + en-us). Partial groundwork is there, just needs locale + language-header fix."),
    (31, 1, "PRIORITY ACTIONS"),
    (32, 1, "1. Sunset arabic.lifeusa.org and donation.lifeusa.org. Confirm 301s are not still pointing to them. Remove their DNS records."),
    (33, 1, "2. Decide on staff.lifeusa.org: move behind authentication, add `Disallow: /` to its robots.txt, AND add a sitewide `noindex` header. Ideally relocate to an internal-only domain off the main brand DNS."),
    (34, 1, "3. Open a Wix support ticket on the `content-language` HTTP header inconsistency. Provide the curl evidence (5 different page samples returning 5 different language codes)."),
    (35, 1, "4. Change hreflang `ar-jo` to `ar` (or to a multi-region pattern with x-default fallback) in Wix Multilingual settings."),
    (36, 1, "5. Audit the 122 noindex'd URLs against the live navigation. Remove noindex from any program, donate, or content page that should rank."),
    (37, 1, "6. Image program: compress the top 50 traffic pages' hero images, then site-wide. Add explicit width/height to every Wix image to fix CLS."),
    (38, 1, "7. Schema upgrade: replace ProfessionalService with NGO type. Add DonateAction. Add Article schema on the 441 blog posts (Wix has a JSON-LD field per post)."),
    (39, 1, "8. Content hygiene pass: meta descriptions on 144 pages, unique titles vs H1 on 97 pages, alt text on 471 images. Prioritize the 33 ranking pages first."),
    (40, 1, "9. Sitemap cleanup: remove /no, /copy-of-freewill-landing-page, /b2s, /fb-intl. Investigate the 4 robots-blocked URLs."),
    (41, 1, "10. Add donor-trust security headers (CSP, X-Frame-Options, Referrer-Policy). Wix has limited control here, may require Wix Studio or a CDN layer; fully achievable on donate. and donation. via Cloudflare."),
    (43, 1, "WIX-SPECIFIC ACTIONS"),
    (44, 1, "* Wix is the CMS (verified via meta generator + Pepyaka server + parastorage/wixstatic CDN). Some controls (server headers, robots.txt template) are platform-managed."),
    (45, 1, "* content-language HTTP header bug: not user-configurable in Wix; requires a support ticket. Document the per-URL evidence first."),
    (46, 1, "* Image width/height: Wix injects images server-side; the `images_missing_size_attributes` issue is likely a Wix template-version issue. Test on a few pages. Newer Wix templates may already include them."),
    (47, 1, "* Hreflang locales: configurable in Wix Multilingual settings (Site Languages > advanced)."),
    (48, 1, "* Schema: Wix supports custom JSON-LD per page via SEO Settings > Advanced > Custom Meta Tags. Upgrade homepage to NGO and add Article on blog posts via the Wix Blog SEO panel."),
    (49, 1, "* Robots.txt: Wix auto-generates. Edit via SEO Tools > Robots.txt Editor. Has a `Crawl-delay: 10` for AhrefsBot which is fine."),
]
for row, col, val in EXEC:
    safe_set(ws, row, col, val)

for r, _, _ in EXEC:
    if r in (1, 7, 14, 31, 43):
        cell = ws.cell(row=r, column=1)
        cell.font = Font(bold=True, size=12)

print("Executive Summary populated.")

# ============================================================================
# TAB 2: Audit Checklist
# ============================================================================
ws = wb['Audit Checklist']

clear_row(ws, 1, 1, 12)
clear_row(ws, 2, 1, 12)
safe_set(ws, 1, 2, "Website being checked:\n\nhttps://www.lifeusa.org/")
safe_set(ws, 1, 4, "robots.txt:\nUser-agent: *\nAllow: /\nDisallow: *?lightbox=\nSitemap: https://www.lifeusa.org/sitemap.xml\nSitemap: https://www.lifeusa.org/ar_ar-sitemap.xml\n(auto-generated by Wix)")

ROWS = {
    6:  ("Needs input", "NEEDS INPUT: GSC access requested via WhatsApp 2026-05-07, awaiting grant for sgamal2593@gmail.com", "Critical"),
    7:  ("Needs input", "NEEDS INPUT: GA4 access not yet requested. Will ask in kickoff.", "Critical"),
    8:  ("Ready",       "See 'Current Page templates' tab. 8 templates documented.", "Moderate"),
    9:  ("Ready",       "Wix.com Website Builder confirmed via meta generator + Pepyaka server + static.wixstatic.com CDN. Donate. on Cloudflare/TouchPoints. Donation. on Cloudflare/Givecloud. Staff. on Cloudflare/WordPress. Arabic. dead in DNS.", "Low"),
    10: ("Needs input", "NEEDS INPUT: ask client about migrations, especially when arabic.lifeusa.org was decommissioned and why donation. duplicates donate.", "Moderate"),
    12: ("Attention",   "Wix is hybrid (some SSR, some CSR). Homepage HTML lacks visible H1 in raw response, likely JS-injected. Not blocking but a signal-strength issue.", "Moderate"),
    13: ("Needs input", "NEEDS INPUT: Visual JS-disabled test required (Wix nav typically progressive but verify on this template).", "Moderate"),
    14: ("Needs input", "NEEDS INPUT: Visual JS-disabled test required.", "Moderate"),
    15: ("NA",          "Wix blog pagination. Verify behavior. Typically progressive in newer Wix templates.", "Low"),
    16: ("Pass",        "Verified. Anchor tags present in raw HTML.", "Low"),
    17: ("Pass",        "Verified. <img> tags with wixstatic.com sources.", "Low"),
    18: ("Needs input", "NEEDS INPUT: Visual test for cookie/consent banners blocking content.", "Low"),
    20: ("NA",          "No faceted product filtering on the site.", "Low"),
    21: ("Attention",   "Multiple landing pages for similar campaigns: /gaza vs /gazaprojects, /ramadan-2025 vs /ramadan-orphan-2025. Risk of cannibalization.", "Moderate"),
    22: ("Needs input", "NEEDS INPUT: confirm canonicals on those overlapping campaign pages.", "Moderate"),
    23: ("NA",          "No e-commerce variations.", "Low"),
    24: ("NA",          "No product page variations.", "Low"),
    26: ("Attention",   "5 internal HTTP URLs detected by Screaming Frog. See security_http_urls.csv.", "High"),
    27: ("Fail",        "5 mixed-content URLs (HTTP resources on HTTPS pages). Donor-trust risk. See security_http_urls_inlinks.csv.", "High"),
    28: ("Needs input", "NEEDS INPUT: GSC > Mobile Usability needed. Wix is responsive by default.", "Moderate"),
    29: ("Needs input", "NEEDS INPUT: <head> validity check. No noscript/invalid tags found in homepage spot-check.", "Low"),
    30: ("Fail",        "37 pages missing H1 (14%). 9 pages with multiple H1 (3%). 42 H1 over 70 chars (16%). 97 page titles identical to H1 (37%). Wix template defaults dominating. See h1_missing.csv, h1_multiple.csv, page_titles_same_as_h1.csv.", "Critical"),
    31: ("Attention",   "WebSite + ProfessionalService schema on homepage. Missing: NGO (more appropriate for charity), DonateAction, Article on 441 blog posts, BreadcrumbList sitewide, FAQPage on policies.", "High"),
    32: ("Needs input", "NEEDS INPUT: GSC > Enhancements needed for per-template validation.", "Moderate"),
    33: ("Fail",        "Significant: NGO type, DonateAction, FAQPage, BreadcrumbList, Article schema all missing or generic.", "High"),
    35: ("Pass",        "Found at https://www.lifeusa.org/robots.txt. Wix auto-generated.", "Critical"),
    36: ("Pass",        "Standard location.", "Critical"),
    37: ("Pass",        "200 OK confirmed via curl.", "Critical"),
    38: ("Pass",        "Sitemap declared: /sitemap.xml + /ar_ar-sitemap.xml", "Low"),
    39: ("Fail",        "Subdomain robots.txt is fragmented. staff.lifeusa.org has its own WordPress robots.txt. arabic.lifeusa.org dead. donate. and donation. on different platforms (TouchPoints vs Givecloud), not coordinated.", "Critical"),
    40: ("Attention",   "Robots blocks `*?lightbox=` (Wix lightbox query, fine). 4 internal URLs blocked by robots. See response_codes_internal_blocked_by_robots_txt.csv. Investigate.", "Critical"),
    41: ("Needs input", "NEEDS INPUT: GSC > Pages > Indexed but blocked by robots.", "Critical"),
    42: ("Pass",        "Wix-managed CSS not blocked.", "Critical"),
    43: ("Pass",        "Wix-managed JS not blocked.", "Critical"),
    44: ("Needs input", "NEEDS INPUT: GSC > Settings > Crawl Stats.", "Critical"),
    45: ("Needs input", "NEEDS INPUT: GSC > Crawl Stats > Response codes.", "Critical"),
    46: ("Attention",   "Screaming Frog flagged 4 internal 4xx; minor on this site. Verify in GSC.", "Critical"),
    47: ("Attention",   "18 internal redirects detected. Mostly single-hop, but worth chain analysis. See response_codes_internal_redirection_(3xx).csv.", "Critical"),
    49: ("Pass",        "33 organic pages indexed per Ahrefs (May 2026). 1,069 monthly organic traffic. Indexation working overall.", "Critical"),
    50: ("Fail",        "122 URLs (35%) carry noindex directives. Needs urgent audit. Likely accidental noindex on key pages. See directives_noindex.csv.", "Critical"),
    51: ("Attention",   "15 pages missing canonicals; rest are self-referencing (Wix default). See canonicals_missing.csv.", "Critical"),
    52: ("Needs input", "NEEDS INPUT: GSC > Pages > Indexed > Inspect.", "Critical"),
    53: ("Attention",   "4 internal 4xx URLs detected via SF. See response_codes_internal_client_error_(4xx).csv.", "Critical"),
    54: ("Needs input", "NEEDS INPUT: GSC > Pages > Not Indexed > Soft 404.", "Critical"),
    55: ("Needs input", "NEEDS INPUT: visual check on 404 page implementation.", "Moderate"),
    56: ("Attention",   "18 internal redirects, mostly single-hop. See response_codes_internal_redirection_(3xx).csv.", "Moderate"),
    57: ("Needs input", "NEEDS INPUT: Screaming Frog rendered crawl needed to confirm.", "Moderate"),
    58: ("Pass",        "lifeusa.org redirects to https://www.lifeusa.org/ via 301. Verified via curl.", "Critical"),
    59: ("Pass",        "All HTTPS verified. HSTS enabled.", "Critical"),
    60: ("Attention",   "/ar/ redirects to /ar (no trailing slash). Generally consistent on Wix.", "Moderate"),
    61: ("Fail",        "staff.lifeusa.org is a separate publicly reachable WordPress install. May be indexable. Risks indexation of internal content.", "Critical"),
    62: ("Needs input", "NEEDS INPUT: site:lifeusa.org inurl:?s= check.", "Moderate"),
    63: ("Pass",        "https://www.lifeusa.org/sitemap.xml is a Wix-generated sitemap index with 4 child sitemaps + 2 ar_ar sitemaps.", "Critical"),
    64: ("Pass",        "loc + lastmod present on all entries.", "Critical"),
    65: ("Needs input", "NEEDS INPUT: confirm in GSC.", "Critical"),
    66: ("Needs input", "NEEDS INPUT: GSC > Sitemaps.", "Critical"),
    67: ("Attention",   "Sitemap contains template leftover URLs: /no, /copy-of-freewill-landing-page, /b2s, /fb-intl. Likely Wix scaffolding never cleaned up.", "Critical"),
    68: ("Needs input", "NEEDS INPUT: GSC.", "Critical"),
    69: ("Needs input", "NEEDS INPUT: cross-check noindex list with robots disallow list.", "Moderate"),
    71: ("Attention",   "Programs section deep. Arabic content at /ar/ separate from English; cross-language linking weak. Donations split across two platforms (donate. + donation.) confuses navigation.", "Critical"),
    72: ("Attention",   "9 URLs uppercase. 22 over 115 chars. 9 with parameters. 2 with underscores. Mostly Wix-managed paths.", "Moderate"),
    73: ("Attention",   "9 URLs with parameters. 3 with GA tracking params (utm/_ga). See url_ga_tracking_parameters.csv.", "Moderate"),
    74: ("Needs input", "NEEDS INPUT: spot-check after parameter URL list.", "Moderate"),
    75: ("Needs input", "NEEDS INPUT: full SF crawl depth report.", "Moderate"),
    76: ("Attention",   "Cross-language linking opportunity. Arabic article discoverability depends entirely on hreflang signal which has the ar-jo locale issue.", "High"),
    77: ("Attention",   "8 internal links lack anchor text. 2 use non-descriptive text (click here / learn more). See links_internal_outlinks_with_no_anchor_text.csv.", "Moderate"),
    78: ("Attention",   "3 internal links carry GA tracking params. Strips session attribution. Remove utm= on internal links.", "Moderate"),
    79: ("Needs input", "NEEDS INPUT: full link graph analysis.", "Moderate"),
    80: ("Needs input", "NEEDS INPUT: orphan analysis from SF crawl.", "Critical"),
    81: ("Attention",   "441 indexed blog posts but limited internal cross-linking from program pages to relevant blog content. Big opportunity to lift the long-tail content into ranking surface.", "Moderate"),
    82: ("Pass",        "Wix blog uses pagination.", "Low"),
    83: ("Needs input", "NEEDS INPUT: confirm pagination URL pattern.", "Low"),
    84: ("Attention",   "3 pagination URLs not in anchor tags per SF. See pagination_pagination_url_not_in_anchor_tag.csv.", "Moderate"),
    86: ("Needs input", "NEEDS INPUT: GSC > Core Web Vitals.", "Critical"),
    87: ("Attention",   "543 images over 100KB (47%). LCP risk on hero images, particularly on Gaza, Programs, donation landing pages.", "Critical"),
    88: ("Needs input", "NEEDS INPUT: GSC > CWV > INP.", "Critical"),
    89: ("Fail",        "1,035 images (90%) missing width/height attributes. Major CLS risk. See images_missing_size_attributes.csv.", "Critical"),
    90: ("Fail",        "543 images over 100KB. 89 alt-text entries over 100 chars. Image format/compression review needed across the site.", "Critical"),
    91: ("Needs input", "NEEDS INPUT: PSI test on top 5 pages.", "Critical"),
    92: ("Needs input", "NEEDS INPUT: PSI 3rd-party-script audit. Wix typically loads many.", "Moderate"),
}

for row, (status, comment, priority) in ROWS.items():
    safe_set(ws, row, 3, status)
    safe_set(ws, row, 4, comment)
    safe_set(ws, row, 5, priority)

print(f"Audit Checklist populated: {len(ROWS)} rows.")

# ============================================================================
# TAB 3: Hints + URL Lists
# ============================================================================
ws = wb['Hints + URL Lists']
clear_sheet_data(ws, header_row=1, max_clear_row=40)

HINTS_HEADER = [
    "Category", "Issue", "% Affected", "# Affected",
    "Severity", "Issue Type", "Reference", "Affected URLs (CSV path)",
]
for i, h in enumerate(HINTS_HEADER, 1):
    safe_set(ws, 1, i, h)

HINTS = [
    ("Subdomain Governance", "arabic.lifeusa.org dead in DNS, donation.lifeusa.org duplicates donate. on Givecloud, staff.lifeusa.org is a public WordPress install", "100%", 4, "1 - Critical", "1 - Issue", "Manual finding (curl + Ahrefs subdomain map)", "Manual finding"),
    ("Indexability", "URLs carrying a noindex directive", "35.36%", 122, "2 - High", "1 - Issue", "https://www.searchenginejournal.com/noindex-tag/", "screaming frog issues/directives_noindex.csv"),
    ("Indexability", "URLs missing a canonical tag", "4.97%", 15, "2 - High", "1 - Issue", "https://developers.google.com/search/docs/crawling-indexing/consolidate-duplicate-urls", "screaming frog issues/canonicals_missing.csv"),
    ("Hreflang/Language", "HTTP content-language header inconsistent across pages (zh-CN, tr-TR, fr-FR observed)", "100% sample", "5 of 5", "1 - Critical", "1 - Issue", "https://developers.google.com/search/docs/specialty/international/localized-versions", "Manual finding (curl headers)"),
    ("Hreflang/Language", "hreflang locale set to ar-jo (Jordanian Arabic) for a global Muslim charity audience", "100%", 1, "2 - High", "1 - Issue", "https://developers.google.com/search/docs/specialty/international/localized-versions", "Manual finding (homepage HTML)"),
    ("Schema/Structured Data", "Generic ProfessionalService schema instead of NGO; Article schema absent on 441 blog posts", "100%", 442, "2 - High", "1 - Issue", "https://schema.org/NGO", "Manual finding (homepage JSON-LD inspection)"),
    ("Page Experience", "Images missing width/height attributes (CLS risk)", "90.47%", 1035, "1 - Critical", "1 - Issue", "https://web.dev/articles/optimize-cls", "screaming frog issues/images_missing_size_attributes.csv"),
    ("Page Experience", "Images over 100KB (LCP/page-speed risk)", "47.47%", 543, "2 - High", "1 - Issue", "https://web.dev/articles/optimize-lcp", "screaming frog issues/images_over_100_kb.csv"),
    ("Accessibility/SEO", "Images missing alt text", "39.42%", 451, "2 - Medium", "1 - Issue", "https://www.w3.org/WAI/tutorials/images/", "screaming frog issues/images_missing_alt_text.csv"),
    ("Accessibility/SEO", "Images missing alt attribute entirely", "1.75%", 20, "2 - Medium", "1 - Issue", "https://www.w3.org/WAI/tutorials/images/", "screaming frog issues/images_missing_alt_attribute.csv"),
    ("Duplicate Content", "Exact duplicate pages", "2.32%", 7, "2 - High", "1 - Issue", "https://developers.google.com/search/docs/crawling-indexing/consolidate-duplicate-urls", "screaming frog issues/content_exact_duplicates.csv"),
    ("Duplicate Content", "Duplicate meta descriptions", "4.53%", 12, "3 - Medium", "1 - Issue", "https://developers.google.com/search/docs/appearance/snippet", "screaming frog issues/meta_description_duplicate.csv"),
    ("On Page", "Page titles identical to H1", "36.60%", 97, "3 - Medium", "2 - Opportunity", "https://developers.google.com/search/docs/appearance/title-link", "screaming frog issues/page_titles_same_as_h1.csv"),
    ("On Page", "Page titles over 60 characters", "27.17%", 72, "3 - Medium", "2 - Opportunity", "https://developers.google.com/search/docs/appearance/title-link", "screaming frog issues/page_titles_over_60_characters.csv"),
    ("On Page", "Meta descriptions missing", "54.34%", 144, "3 - Medium", "2 - Opportunity", "https://developers.google.com/search/docs/appearance/snippet", "screaming frog issues/meta_description_missing.csv"),
    ("On Page", "H1 missing", "13.96%", 37, "3 - Medium", "1 - Issue", "https://developers.google.com/search/docs/appearance/structured-data", "screaming frog issues/h1_missing.csv"),
    ("On Page", "Multiple H1 on same page", "3.40%", 9, "3 - Medium", "1 - Issue", "https://developers.google.com/search/docs/appearance/structured-data", "screaming frog issues/h1_multiple.csv"),
    ("On Page", "H1 over 70 characters", "15.85%", 42, "4 - Low", "2 - Opportunity", "https://developers.google.com/search/docs/appearance/structured-data", "screaming frog issues/h1_over_70_characters.csv"),
    ("Indexability", "Internal pages blocked by robots.txt", "0.21%", 4, "2 - High", "1 - Issue", "https://developers.google.com/search/docs/crawling-indexing/robots/intro", "screaming frog issues/response_codes_internal_blocked_by_robots_txt.csv"),
    ("Indexability", "Internal 4xx errors", "0.21%", 4, "2 - High", "1 - Issue", "https://developers.google.com/search/docs/crawling-indexing/http-network-errors", "screaming frog issues/response_codes_internal_client_error_(4xx).csv"),
    ("Indexability", "Internal redirects (3xx)", "0.97%", 18, "3 - Medium", "1 - Issue", "https://developers.google.com/search/docs/crawling-indexing/301-redirects", "screaming frog issues/response_codes_internal_redirection_(3xx).csv"),
    ("Security", "HTTP URLs (mixed content)", "1.35%", 5, "2 - High", "1 - Issue", "https://web.dev/articles/what-is-mixed-content", "screaming frog issues/security_http_urls.csv"),
    ("Security/Donor Trust", "Missing Content-Security-Policy header", "88.95%", 330, "4 - Low", "1 - Issue", "https://web.dev/articles/strict-csp", "screaming frog issues/security_missing_contentsecuritypolicy_header.csv"),
    ("Security/Donor Trust", "Missing X-Frame-Options header", "88.95%", 330, "4 - Low", "1 - Issue", "https://developer.mozilla.org/en-US/docs/Web/HTTP/Headers/X-Frame-Options", "screaming frog issues/security_missing_xframeoptions_header.csv"),
    ("Security/Donor Trust", "Missing Referrer-Policy header", "92.99%", 345, "4 - Low", "1 - Issue", "https://developer.mozilla.org/en-US/docs/Web/HTTP/Headers/Referrer-Policy", "screaming frog issues/security_missing_secure_referrerpolicy_header.csv"),
    ("Internal Links", "Internal outlinks with no anchor text", "3.02%", 8, "3 - Medium", "2 - Opportunity", "https://developers.google.com/search/docs/crawling-indexing/links-crawlable", "screaming frog issues/links_internal_outlinks_with_no_anchor_text.csv"),
    ("URL Hygiene", "URLs with GA tracking parameters", "0.81%", 3, "3 - Medium", "1 - Issue", "https://support.google.com/analytics/answer/10917952", "screaming frog issues/url_ga_tracking_parameters.csv"),
    ("Sitemap Hygiene", "Template leftover URLs in sitemap (/no, /copy-of-freewill-landing-page, /b2s, /fb-intl)", "manual", 4, "3 - Medium", "1 - Issue", "https://developers.google.com/search/docs/crawling-indexing/sitemaps/build-sitemap", "Manual finding (sitemap.xml inspection)"),
]

for i, row in enumerate(HINTS, start=2):
    for j, val in enumerate(row, start=1):
        safe_set(ws, i, j, val)

for j in range(1, 9):
    ws.cell(row=1, column=j).font = Font(bold=True)

print(f"Hints + URL Lists populated: {len(HINTS)} rows.")

# ============================================================================
# TAB 4: Current Page templates
# ============================================================================
ws = wb['Current Page templates']
clear_sheet_data(ws, header_row=1, max_clear_row=30)

PT_HEADER = ["Page Template", "Example URL (English)", "Example URL (Arabic)", "URL Pattern", "SEO Status", "Comments / Issues"]
for i, h in enumerate(PT_HEADER, 1):
    safe_set(ws, 1, i, h)

TEMPLATES = [
    ("Homepage", "https://www.lifeusa.org/", "https://www.lifeusa.org/ar", "/[/ar]", "Issue", "Title same as H1 issue. content-language HTTP header returns zh-CN (incorrect). Schema is ProfessionalService, should be NGO. hreflang ar-jo set."),
    ("About / Mission", "https://www.lifeusa.org/about", "https://www.lifeusa.org/ar/about", "/about [/ar/about]", "Issue", "content-language HTTP returns tr-TR (Turkish, incorrect). Stale meta description claims '29+ years' (founded 1992, currently 34 years)."),
    ("Programs / What We Do", "https://www.lifeusa.org/programs", "(check /ar/programs)", "/programs", "Issue", "0 organic traffic per Ahrefs. Likely missed schema and weak internal linking. Programs section structure deep."),
    ("Campaign Landing", "https://www.lifeusa.org/gaza", "(check /ar/gaza)", "/[campaign-name]", "Issue", "Top traffic landing (29 organic). content-language HTTP returns fr-FR (incorrect). Risk of cannibalization vs /gazaprojects. Schema not campaign-specific."),
    ("Donation (Primary)", "https://donate.lifeusa.org/donorportal/", "n/a (single language)", "donate.lifeusa.org", "Attention", "Separate platform (TouchPoints/Cloudflare). Hardened security headers (CSP, HSTS). 214 referring domains, but only 4 organic traffic. Donor-intent queries not captured."),
    ("Donation (Duplicate)", "https://donation.lifeusa.org/", "n/a", "donation.lifeusa.org", "Critical", "Second donation platform (Givecloud/Cloudflare). Duplicates donate. functionality. Splits donor data and SEO equity. Recommended: sunset."),
    ("Blog Post", "https://www.lifeusa.org/post/[slug]", "https://www.lifeusa.org/ar/post/[slug]", "/post/[slug]", "Issue", "441 indexed blog posts. Generic Wix blog template. Article schema absent. Few cross-linked from program pages. Top performers: earthquakes, Ramadan, water wells, qurbani. Coincidental rankings, not charity-intent."),
    ("Board Members", "https://www.lifeusa.org/board-members", "(check /ar/board-members)", "/board-members", "Pass", "10 organic traffic, UR 8. Strong page. Could add Person schema for each board member."),
    ("Careers", "https://www.lifeusa.org/careers", "(check /ar/careers)", "/careers", "Attention", "9 organic traffic, page is new. Could add JobPosting schema for each open role to surface in Google Jobs."),
    ("Staff Portal (subdomain)", "https://staff.lifeusa.org/", "n/a", "staff.lifeusa.org", "Critical", "Public WordPress install with no auth gate visible. Risks indexation of internal content. Recommended: move behind auth + Disallow: / + sitewide noindex."),
    ("Arabic Subdomain (sunset)", "https://arabic.lifeusa.org/ (DEAD)", "n/a", "arabic.lifeusa.org", "Critical", "Does not resolve in DNS. 1 referring domain wasted. Sunset/remove DNS records to stop confusing Ahrefs and any external linkers."),
    ("Sitemap Leftover URLs", "/no, /copy-of-freewill-landing-page, /b2s, /fb-intl", "n/a", "various", "Issue", "Wix template scaffolding URLs still in sitemap. Need investigation. Likely 410 or remove from sitemap."),
]

for i, row in enumerate(TEMPLATES, start=2):
    for j, val in enumerate(row, start=1):
        safe_set(ws, i, j, val)

SUMMARY_ROW = len(TEMPLATES) + 4
safe_set(ws, SUMMARY_ROW,     1, "SUMMARY")
safe_set(ws, SUMMARY_ROW + 1, 1, "Total Page Templates Documented:")
safe_set(ws, SUMMARY_ROW + 1, 2, len(TEMPLATES))
safe_set(ws, SUMMARY_ROW + 2, 1, "Critical Issues:")
safe_set(ws, SUMMARY_ROW + 2, 2, "Subdomain sprawl (arabic./donation./staff.); content-language HTTP header bug; ProfessionalService instead of NGO schema")
safe_set(ws, SUMMARY_ROW + 3, 1, "Primary Problem:")
safe_set(ws, SUMMARY_ROW + 3, 2, "Subdomain governance is fragmenting backlinks, donor data, and crawl budget. Wix HTTP language header bug is suppressing Arabic and English brand discoverability.")

ws.cell(row=SUMMARY_ROW, column=1).font = Font(bold=True, size=12)
for j in range(1, 7):
    ws.cell(row=1, column=j).font = Font(bold=True)

print(f"Current Page Templates populated: {len(TEMPLATES)} rows.")

# ============================================================================
# TAB 5: Prioritization pointers (UNTOUCHED)
# ============================================================================
print("Prioritization pointers: UNTOUCHED (legend)")

# ============================================================================
# TAB 6: Expanded Issues Analysis
# ============================================================================
ws = wb['Expanded Issues Analysis']
clear_sheet_data(ws, header_row=1, max_clear_row=70)

EI_HEADER = ["Category", "Issue / Finding", "Affected URLs", "Percentage", "Priority", "Issue Type", "Technical Description (for Dev/Wix Team)", "Recommended Fix"]
for i, h in enumerate(EI_HEADER, 1):
    safe_set(ws, 1, i, h)

EI = [
    ("Subdomain Governance", "Subdomain Sprawl Fragments Equity and Donor Data",
     "4 subdomains", "100%", "Critical", "Issue",
     "Five subdomains are in flight: www. (live), donate. (TouchPoints), donation. (Givecloud, duplicates donate.), arabic. (DEAD in DNS, 1 ref domain wasted), staff. (public WordPress, separate install). Apex 301 to www works. Each non-www subdomain is on different infrastructure with different security postures and different SEO postures.",
     "1. Sunset arabic.lifeusa.org: confirm no live links pointing to it, remove DNS A/CNAME records.\n2. Sunset donation.lifeusa.org: pick one donations platform (donate. on TouchPoints is more SEO-mature with 214 ref domains), 301 donation.* to donate. equivalents.\n3. staff.lifeusa.org: move behind authentication, add Disallow: / + sitewide noindex header, ideally relocate to a non-brand internal domain.\n4. Add Sitemap directive in www.robots.txt that points only to canonical sitemaps."),
    ("Hreflang/Language", "HTTP content-language Header Inconsistent Per Page",
     "Sample: 5 pages tested", "100%", "Critical", "Issue",
     "Wix is sending randomized/incorrect content-language HTTP response headers per page: / returns zh-CN (Chinese), /about returns tr-TR (Turkish), /gaza returns fr-FR (French), /ar returns en (English instead of ar), /post/* returns en. The HTML <html lang> attribute is correct. Google uses these headers as signals; inconsistency suppresses brand visibility especially in English search.",
     "1. Open a Wix support ticket. Provide the curl evidence (5 URL > 5 different language codes).\n2. In the meantime, ask Wix support whether the Multilingual product is generating these headers from the user's browser Accept-Language vs. from the page's declared language. The cache hits suggest the headers are being baked into the cache layer.\n3. Re-test after Wix fix; verify across 10+ pages."),
    ("Hreflang/Language", "hreflang Locale Set to ar-jo for a Global Muslim Charity",
     "All Arabic alternates", "100%", "High", "Issue",
     "Site declares hreflang='ar-jo' (Jordanian Arabic) for the /ar version. The charity's audience includes Muslim donors across Saudi Arabia, Egypt, the Gulf, Morocco, and the Arab diaspora. ar-jo is too narrow.",
     "1. Change hreflang to 'ar' in Wix Multilingual settings.\n2. Confirm the en alternate uses 'en' (or 'en-us' which is also acceptable for a US org).\n3. Keep x-default pointing to the English version."),
    ("Indexability", "122 URLs Carry noindex Directive (Likely Accidental)",
     "122 URLs", "35.36%", "Critical", "Issue",
     "Screaming Frog detected 122 URLs (35% of crawled URLs) carrying a noindex directive. On a charity with only 33 ranking pages, having 35% of the site noindexed suggests accidental noindex on important program/landing pages, OR Wix-managed system pages that Wix already excludes.",
     "1. Open screaming frog issues/directives_noindex.csv.\n2. Cross-reference with Wix > SEO Tools > Indexability per page.\n3. Remove noindex from any program, donation, campaign, or content page that should rank.\n4. Wix system pages (login, thank-you, account) are appropriately noindexed. Skip those."),
    ("Schema/Structured Data", "ProfessionalService Schema Instead of NGO; Article Schema Missing",
     "Homepage + 441 posts", "100%", "High", "Issue",
     "Homepage JSON-LD declares schema.org/ProfessionalService. For a registered nonprofit charity, the more semantically appropriate type is schema.org/NGO (subclass of Organization). DonateAction is also missing. Article schema is absent on 441 indexed blog posts. This is the largest schema opportunity on the site.",
     "1. In Wix > SEO Settings > Advanced > Custom Meta Tags on the homepage, replace the auto-generated ProfessionalService block with NGO including: name, alternateName (Arabic), url, logo, description, founder, foundingDate (1992), address, taxID/EIN, areaServed, knowsAbout, nonprofitStatus, sameAs (social profiles), contactPoint, hasMap.\n2. Add DonateAction schema linking to donate.lifeusa.org.\n3. For each blog post: enable Wix Blog SEO panel JSON-LD with Article (or NewsArticle if breaking news), including author, datePublished, image, publisher.\n4. Add BreadcrumbList sitewide via the Wix breadcrumb component."),
    ("Page Experience", "Images Missing width/height Attributes (CLS Risk)",
     "1,035 images", "90.47%", "Critical", "Issue",
     "1,035 images (90% of crawled images) lack width/height attributes in the HTML. Browsers cannot reserve space before the image loads, causing Cumulative Layout Shift. CLS over 0.1 fails Core Web Vitals.",
     "1. This is largely a Wix template-version issue. Newer Wix templates include explicit dimensions; older templates do not.\n2. In Wix Editor, re-upload key hero images using the latest image component to force the new template behavior.\n3. As a sitewide fallback, contact Wix support to enable explicit dimensions on the rendered <img> tag.\n4. Verify on top 5 pages with Chrome DevTools Performance panel."),
    ("Page Experience", "543 Images Over 100KB (LCP Risk)",
     "543 images", "47.47%", "Critical", "Issue",
     "543 images exceed 100KB. Hero images on Gaza, programs, and donation landing pages are most impacted by Largest Contentful Paint scoring. Page-speed degradation also impacts donor conversion.",
     "1. Audit images on top 5 traffic pages first.\n2. Replace JPG/PNG hero images with WebP or AVIF (Wix supports both).\n3. Use Wix's image compression option in the Editor.\n4. Lazy-load below-the-fold images.\n5. Re-test with PSI after each round."),
    ("Accessibility/SEO", "451 Images Missing Alt Text + 20 Missing alt Attribute",
     "471 images", "41.17%", "High", "Issue",
     "451 images have an alt attribute but the alt text is empty. 20 images are missing the alt attribute entirely. This directly maps to the client's pain point: Arabic articles' images do not surface in English image search because the images have no descriptive text in any language.",
     "1. Bulk-update alt text via Wix Editor on the top 50 traffic pages first.\n2. For Arabic articles: add bilingual alt text where possible (Arabic primary, with English transliteration in parens for proper nouns).\n3. Decorative images: explicit alt='' is acceptable.\n4. Train Tasneem and content team: every image upload requires alt text before publish."),
    ("Duplicate Content", "Exact Duplicate Pages",
     "7 URLs", "2.32%", "High", "Issue",
     "7 pages have identical HTML hashes per Screaming Frog. This splits PageRank and confuses ranking signals.",
     "1. Identify pairs in screaming frog issues/content_exact_duplicates.csv.\n2. Pick canonical version of each pair.\n3. 301 the duplicate to the canonical, OR add canonical tag pointing to the canonical.\n4. Update internal links to point to the canonical only."),
    ("Duplicate Content", "12 Duplicate Meta Descriptions",
     "12 URLs", "4.53%", "Medium", "Issue",
     "12 URLs share meta description text. Wix may be auto-filling from the first paragraph on similar campaign pages (e.g., overlapping Ramadan or Gaza variants).",
     "1. See screaming frog issues/meta_description_duplicate.csv.\n2. Write unique 150-character meta descriptions per page in Wix > SEO Settings.\n3. Prioritize ranking pages."),
    ("On Page", "Page Titles Identical to H1",
     "97 URLs", "36.60%", "Medium", "Opportunity",
     "97 pages (37%) have title tags matching H1 exactly. This is a Wix template default. Not technically broken but a missed opportunity to target related keywords or front-load donor-intent keywords in the title.",
     "1. For top 33 ranking pages, write distinct titles vs H1: H1 = topic, Title = topic + brand + intent qualifier.\n2. Example: H1 'Gaza Emergency Relief'; Title 'Gaza Emergency Relief Donations | Life for Relief and Development'.\n3. Edit per page in Wix > SEO Settings."),
    ("On Page", "144 Pages Missing Meta Description",
     "144 URLs", "54.34%", "Medium", "Opportunity",
     "More than half the indexed pages have no meta description. Google generates a snippet from page content, often poorly for charity pages where mission language is generic.",
     "1. Write meta descriptions on the 33 ranking pages first (highest CTR impact).\n2. 130-155 characters, action-oriented, include donor-intent keyword + brand.\n3. Bulk-edit via Wix Editor or use a CSV approach if Wix Multilingual supports it."),
    ("On Page", "37 Pages Missing H1; 9 Pages with Multiple H1",
     "46 URLs", "17.36%", "Medium", "Issue",
     "37 pages lack an H1; 9 pages have multiple H1s. Strong on-page ranking signal degraded.",
     "1. Add a single H1 per page in Wix Editor. Usually the page title or an obvious top heading.\n2. For multi-H1 pages: convert all but the primary to H2.\n3. Verify the H1 contains the primary topic keyword."),
    ("Indexability", "Internal 4xx Errors",
     "4 URLs", "0.21%", "High", "Issue",
     "4 internal URLs return 4xx (broken links). Inlinks export shows where they are linked from.",
     "1. See screaming frog issues/response_codes_internal_client_error_(4xx).csv + _inlinks.csv.\n2. Update or remove the broken links at source.\n3. If the destination should exist, restore the page or 301 to the new location."),
    ("Indexability", "4 Internal URLs Blocked by robots.txt",
     "4 URLs", "0.21%", "High", "Issue",
     "robots.txt is auto-generated by Wix. The Disallow rules only block `*?lightbox=` (a Wix UI artifact) yet 4 internal URLs are blocked. Likely a query-string mismatch or a Wix system path.",
     "1. See screaming frog issues/response_codes_internal_blocked_by_robots_txt.csv.\n2. Verify each URL. Confirm if intentional (Wix system paths) or accidental.\n3. If accidental, edit robots.txt in Wix > SEO Tools."),
    ("Security", "5 HTTP URLs (Mixed Content)",
     "5 URLs", "1.35%", "High", "Issue",
     "5 URLs in the crawl are HTTP rather than HTTPS, creating mixed-content warnings on a charity site that needs to project donor trust.",
     "1. See screaming frog issues/security_http_urls.csv.\n2. Update internal links to HTTPS versions.\n3. If they are external resources (image hotlinks), replace with HTTPS sources."),
    ("Security/Donor Trust", "Missing CSP, X-Frame-Options, Referrer-Policy Headers",
     "330-345 URLs", "89-93%", "Low", "Issue",
     "Donor-trust security headers are absent on the Wix-hosted main site. Wix has limited customer control over response headers. donate. and donation. on Cloudflare have CSP correctly set.",
     "1. Open a Wix support ticket requesting CSP, X-Frame-Options (SAMEORIGIN), and Referrer-Policy (strict-origin-when-cross-origin) defaults on the site.\n2. Wix Studio (their newer product) may allow custom response headers. Explore migration if security posture is a priority.\n3. The donations subdomains already have CSP; verify and document for transparency."),
    ("Sitemap Hygiene", "Template Leftover URLs in Sitemap",
     "4 URLs", "manual", "Medium", "Issue",
     "Sitemap entries /no, /copy-of-freewill-landing-page, /b2s, /fb-intl appear to be Wix scaffolding URLs. They consume crawl budget and signal poor site hygiene.",
     "1. Verify each URL in browser. Likely 410-able.\n2. In Wix > SEO Tools, set these pages to noindex + remove from sitemap, OR delete them entirely.\n3. Re-submit sitemap to GSC after cleanup."),
    ("Internal Linking", "Donations Architecture Splits Equity",
     "donate. + donation.", "n/a", "High", "Issue",
     "donate.lifeusa.org has 214 referring domains and 4 organic traffic. donation.lifeusa.org has 4 referring domains. Together they fragment donor links and donor analytics, with no clear primary CTA from www.",
     "1. Pick donate. as primary (more authoritative).\n2. 301 donation.* to the equivalent donate.* path where possible, or to donate.lifeusa.org/donorportal/.\n3. Update all www. CTAs to point to donate. consistently.\n4. Pull both donation reports. Confirm equivalence before migration."),
    ("Content Strategy", "Top Organic Pages Are Coincidental, Not Charity-Intent",
     "33 ranking pages", "n/a", "Medium", "Opportunity",
     "Top organic blog posts: earthquakes, Ramadan, water wells, qurbani, raising empathic kids. These rank on informational queries, not donor-intent queries (e.g. 'donate to gaza relief', 'ramadan zakat charity', 'orphan sponsorship'). The site is winning trickle traffic but not converting it.",
     "1. Add donor-intent CTAs and program links on every blog post (currently weak).\n2. Build content cluster around each program (Gaza, Orphans, Water, Hidayah) with internal linking from the existing trickle posts.\n3. Add Article schema with author + organization to lift E-E-A-T signals."),
]

for i, row in enumerate(EI, start=2):
    for j, val in enumerate(row, start=1):
        safe_set(ws, i, j, val)

for j in range(1, 9):
    ws.cell(row=1, column=j).font = Font(bold=True)

print(f"Expanded Issues Analysis populated: {len(EI)} rows.")

wb.save(DST)
print(f"\nSaved: {DST}")
print(f"File size: {os.path.getsize(DST):,} bytes")
